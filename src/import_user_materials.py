"""Import user-provided product links, job postings, and annual-report files.

No field is invented. Blank company/platform cells in the recruitment workbook
are forward-filled only within the original worksheet, reflecting its visual
"same as above" convention. The original workbook remains untouched.
"""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from common import ensure_directories, repo_path, sha256_file


DEFAULT_PRODUCTS = Path("/Users/emi/Desktop/未命名文件夹/小学期/产品页链接汇总.xlsx")
DEFAULT_JOBS = Path("/Users/emi/Desktop/未命名文件夹/小学期/招聘信息整理不完全.xlsx")
DEFAULT_REPORTS = Path("/Users/emi/Desktop/未命名文件夹/小学期/年报")

COMPANY_ALIASES = {"东方财富集团": "东方财富", "东方财富证券": "东方财富"}
EXPECTED_JOB_COLUMNS = ["序号", "公司名称", "检索平台", "职位名称", "薪资范围", "工作地点", "技能要求（硬技能+软技能）", "学历要求", "经验要求", "AI相关技术栈要求"]


def compact(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def import_products(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        company_label = compact(sheet.cell(row_number, 1).value)
        title = compact(sheet.cell(row_number, 2).value)
        url = sheet.cell(row_number, 2).hyperlink.target if sheet.cell(row_number, 2).hyperlink else title
        company_name = re.sub(r"^\d+\.\s*", "", company_label)
        rows.append({"company_name": company_name, "product_page_title": title, "product_url": compact(url), "source_workbook": path.name, "source_row": row_number, "access_date": date.today().isoformat(), "status": "user_provided_link"})
    return pd.DataFrame(rows)


def import_jobs(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True)
    records = []
    for sheet in workbook.worksheets:
        header = [compact(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        indices = {name: header.index(name) for name in EXPECTED_JOB_COLUMNS if name in header}
        missing = set(EXPECTED_JOB_COLUMNS) - set(indices)
        if missing:
            raise ValueError(f"{sheet.title} is missing expected columns: {sorted(missing)}")
        previous_company = ""
        previous_platform = ""
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            values = {name: compact(row[index]) for name, index in indices.items()}
            if values["公司名称"]:
                previous_company = COMPANY_ALIASES.get(values["公司名称"], values["公司名称"])
            if values["检索平台"]:
                previous_platform = values["检索平台"]
            if not values["职位名称"]:
                continue
            records.append({"company_name": previous_company, "platform": previous_platform, "job_title": values["职位名称"], "salary_raw": values["薪资范围"], "location_raw": values["工作地点"], "skills_raw": values["技能要求（硬技能+软技能）"], "education_raw": values["学历要求"], "experience_raw": values["经验要求"], "ai_tech_stack_raw": values["AI相关技术栈要求"], "source_workbook": path.name, "source_sheet": sheet.title, "source_row": row_number, "import_date": date.today().isoformat()})
    frame = pd.DataFrame(records)
    sample_pool = pd.read_csv(repo_path("data", "reference", "final_sample_pool.csv"), dtype=str, keep_default_na=False)
    company_codes = dict(zip(sample_pool["company_name"], sample_pool["company_code"]))
    frame["company_code"] = frame["company_name"].map(company_codes).fillna("")
    frame["job_description_raw"] = (
        "技能要求：" + frame["skills_raw"]
        + "\n学历要求：" + frame["education_raw"]
        + "\n经验要求：" + frame["experience_raw"]
        + "\nAI相关技术栈要求：" + frame["ai_tech_stack_raw"]
    )
    frame["job_id"] = [f"JOB_{index:04d}" for index in range(1, len(frame) + 1)]
    frame["job_url"] = ""
    frame["publish_time_raw"] = ""
    frame["crawl_date"] = frame["import_date"]
    frame["source_id"] = "user_recruitment_workbook"
    return frame


def copy_reports(directory: Path) -> pd.DataFrame:
    output = repo_path("data", "raw", "annual_reports")
    ensure_directories([output])
    rows = []
    report_files = sorted(path for path in directory.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")
    for source in report_files:
        destination = output / source.name
        if not destination.exists() or sha256_file(destination) != sha256_file(source):
            shutil.copy2(source, destination)
        rows.append({"source_filename": source.name, "local_path": str(destination.relative_to(repo_path())), "sha256": sha256_file(destination), "status": "user_provided_report"})
    return pd.DataFrame(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--products", type=Path, default=DEFAULT_PRODUCTS)
    parser.add_argument("--jobs", type=Path, default=DEFAULT_JOBS)
    parser.add_argument("--reports-dir", type=Path, default=DEFAULT_REPORTS)
    args = parser.parse_args()
    for path in [args.products, args.jobs, args.reports_dir]:
        if not path.exists():
            raise FileNotFoundError(f"User-provided input not found: {path}")
    products = import_products(args.products)
    jobs = import_jobs(args.jobs)
    reports = copy_reports(args.reports_dir)
    processed = repo_path("data", "processed")
    reference = repo_path("data", "reference")
    ensure_directories([processed, reference])
    products.to_csv(reference / "product_pages.csv", index=False, encoding="utf-8")
    jobs.to_csv(repo_path("data", "raw", "jobs", "jobs_raw.csv"), index=False, encoding="utf-8")
    reports.to_csv(reference / "user_provided_annual_reports.csv", index=False, encoding="utf-8")
    print(f"Imported {len(products)} product links, {len(jobs)} job records, and {len(reports)} annual reports.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
